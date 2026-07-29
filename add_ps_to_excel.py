#!/usr/bin/env python3
"""Physical Security 13개 제품을 엑셀 템플릿에 추가"""

import openpyxl
from config import NOCODB_URL, NOCODB_API_TOKEN, NOCODB_TABLE_ID
from nocodb_client import NocoDBTable

# Physical Security 13개 제품 (재고 있는 제품들)
PS_PRODUCTS = [
    # 카메라 5개
    ("UVC-G6-Pro-360", "g6-pro-360"),
    ("UVC-AI-PTZ", "ai-ptz-industrial"),
    ("UVC-G5-Turret-Ultra", "g5-turret-ultra"),
    ("UVC-G6-Dome", "g6-dome"),
    ("UVC-AI-THETA-HUB", "ai-theta"),
    # 센서 3개
    ("UP-Sense", "all-in-one-sensor"),
    ("USL-GlassBreak", "glass-break-sensor"),
    ("USL-Motion", "motion-sensor"),
    # 녹화/컨트롤 2개
    ("UNVR-Instant", "nvr-instant"),
    ("UCK-G2-SSD", "cloudkey-plus"),
    # 기타 3개
    ("UP-AI-Horn-Speaker", "ai-horn-speaker"),
    ("USL-GATEWAY", "superlink-gateway"),
    ("UP-FloodLight", "floodlight"),
]

def main():
    # NocoDB에서 데이터 가져오기
    table = NocoDBTable(NOCODB_URL, NOCODB_API_TOKEN, NOCODB_TABLE_ID)
    records = table.all()

    # Model Number로 레코드 찾기
    model_to_record = {}
    for r in records:
        model = r["fields"].get("Model Number")
        if model:
            model_to_record[model] = r["fields"]

    # 엑셀 파일 열기
    wb = openpyxl.load_workbook('naver_상품등록_템플릿.xlsx')
    ws = wb['상품데이터']

    start_row = ws.max_row + 1
    print(f"75행부터 Physical Security 제품 추가 시작 (현재 {ws.max_row}행까지 있음)\n")

    # Physical Security 카테고리 ID
    # 네트워크장비 > 보안장비 또는 적절한 카테고리
    CATEGORY_NAME = "네트워크장비>보안장비"
    LEAF_CATEGORY_ID = 50001625  # 임시 - 실제 카테고리 ID 필요

    current_row = start_row
    for model_number, page_slug in PS_PRODUCTS:
        fields = model_to_record.get(model_number)
        if not fields:
            print(f"⚠️  {model_number} - NocoDB에서 찾을 수 없음")
            continue

        sku = fields.get("SKU", "")
        final_price = fields.get("최종가격", 0)
        stock = fields.get("In_Stock", 0)

        # 이미지 폴더명은 SKU 기반
        image_folder = sku.replace(" ", "_")

        # 행 데이터 작성
        ws.cell(current_row, 1).value = sku  # 영문상품명
        ws.cell(current_row, 2).value = f"유니파이 {sku}"  # 한글상품명
        ws.cell(current_row, 3).value = CATEGORY_NAME
        ws.cell(current_row, 4).value = LEAF_CATEGORY_ID
        ws.cell(current_row, 5).value = final_price  # 판매가
        ws.cell(current_row, 6).value = final_price  # 할인가 (동일)
        ws.cell(current_row, 7).value = image_folder  # 제품이미지_폴더명
        ws.cell(current_row, 8).value = page_slug  # 상세페이지_폴더명
        ws.cell(current_row, 9).value = None  # 옵션명
        ws.cell(current_row, 10).value = None  # 옵션값
        ws.cell(current_row, 11).value = max(1, stock)  # 재고수량 (0이면 1로)
        ws.cell(current_row, 12).value = "FREE"  # 배송비타입
        ws.cell(current_row, 13).value = 0  # 배송비금액
        ws.cell(current_row, 14).value = "Physical Security"  # 비고

        print(f"✓ {current_row}행: {sku} (가격: {final_price:,}원, 재고: {stock})")
        current_row += 1

    # 저장
    wb.save('naver_상품등록_템플릿.xlsx')
    print(f"\n총 {current_row - start_row}개 제품 추가 완료")
    print(f"엑셀 저장: naver_상품등록_템플릿.xlsx (현재 {ws.max_row}행)")

if __name__ == "__main__":
    main()
