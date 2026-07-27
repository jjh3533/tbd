"""
엑셀 템플릿(naver_상품등록_템플릿.xlsx)을 읽어 네이버 스마트스토어에 상품을 일괄 등록한다.

사용법:
    python main.py --file naver_상품등록_템플릿.xlsx --dry-run   # 실제 등록 없이 확인만
    python main.py --file naver_상품등록_템플릿.xlsx            # 실제 등록 (이미 등록된 상품은 자동으로 건너뜀)
    python main.py --force                                     # 이미 등록된 상품도 다시 등록 (중복 생성 주의)

사전 준비:
    1. config.py 에 CLIENT_ID / CLIENT_SECRET / 주소록 ID / PRODUCT_IMAGES_DIR / PRODUCT_PAGES_DIR 입력
    2. pip install -r requirements.txt
    3. 엑셀 템플릿의 leafCategoryId 를 category_lookup.py 로 조회해서 채워넣기
    4. 엑셀의 '제품이미지_폴더명'/'상세페이지_폴더명'에 Product Images / Product Pages 안의
       실제 폴더명을 입력 (예: "UCG Ultra", "ucg-ultra")
    5. --dry-run 으로 먼저 실행해서 생성되는 JSON이 의도한 대로인지 확인

등록 이력은 registered_log.json에 기록되며, 같은 영문상품명으로 이미 등록에 성공한 행은
기본적으로 다시 등록하지 않고 건너뛴다 (중복 상품 생성 방지).
"""
import argparse
import json
import os
import sys

import openpyxl
import requests

import naver_config as config
from auth import get_bearer_token
from image_uploader import (
    list_detail_page_images,
    list_optional_images,
    pick_representative_image,
    resolve_detail_page_folder,
    resolve_product_image_folder,
    upload_with_retry,
)
from product_builder import build_product_payload

REQUIRED_COLUMNS = [
    "영문상품명", "한글상품명", "카테고리(대분류>중분류)", "leafCategoryId",
    "판매가", "제품이미지_폴더명", "상세페이지_폴더명",
    "재고수량", "배송비타입",
]

LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "registered_log.json")


def load_log() -> dict:
    if not os.path.exists(LOG_PATH):
        return {}
    with open(LOG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_log(log: dict):
    with open(LOG_PATH, "w", encoding="utf-8") as f:
        json.dump(log, f, ensure_ascii=False, indent=2)


def read_rows(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["상품데이터"]
    headers = [c.value for c in ws[1]]

    rows = []
    for excel_row in ws.iter_rows(min_row=2, values_only=False):
        values = [c.value for c in excel_row]
        if not any(values):
            continue
        row = dict(zip(headers, values))
        if not row.get("영문상품명"):
            continue
        missing = [col for col in REQUIRED_COLUMNS if not row.get(col) and row.get(col) != 0]
        if missing:
            print(f"[건너뜀] '{row.get('영문상품명')}' - 필수 항목 누락: {missing}")
            continue
        rows.append(row)
    return rows


def register_one(token: str, row: dict, dry_run: bool):
    name = row["영문상품명"]
    print(f"\n=== {name} ===")

    image_folder = resolve_product_image_folder(str(row["제품이미지_폴더명"]).strip())
    detail_folder = resolve_detail_page_folder(str(row["상세페이지_폴더명"]).strip())

    representative_path = pick_representative_image(image_folder)
    if not representative_path:
        raise RuntimeError(f"대표이미지를 찾을 수 없음: {image_folder}")

    optional_paths = list_optional_images(image_folder, exclude_path=representative_path)
    detail_paths = list_detail_page_images(detail_folder)
    if not detail_paths:
        raise RuntimeError(f"상세페이지 이미지를 찾을 수 없음: {detail_folder}")

    print(f"  대표이미지: {representative_path}")
    print(f"  추가이미지 {len(optional_paths)}장 (최대 {config.MAX_OPTIONAL_IMAGES}장까지만 사용)")
    print(f"  상세페이지 이미지 {len(detail_paths)}장")

    if dry_run:
        image_urls = {
            "representative": f"(dry-run) {representative_path}",
            "optional": [f"(dry-run) {p}" for p in optional_paths],
            "detail": [f"(dry-run) {p}" for p in detail_paths],
        }
    else:
        rep_url = upload_with_retry(token, [representative_path])[0]

        opt_urls = []
        if optional_paths:
            opt_urls = upload_with_retry(token, optional_paths)

        detail_urls = upload_with_retry(token, detail_paths)

        image_urls = {"representative": rep_url, "optional": opt_urls, "detail": detail_urls}

    payload = build_product_payload(row, image_urls)

    if dry_run:
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return None

    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    resp = requests.post(config.PRODUCT_API_URL, headers=headers, json=payload, timeout=30)

    if resp.status_code >= 300:
        print(f"  [실패] {resp.status_code}: {resp.text}")
        return None

    result = resp.json()
    print(
        f"  [성공] originProductNo={result.get('originProductNo')} "
        f"smartstoreChannelProductNo={result.get('smartstoreChannelProductNo')}"
    )
    return result


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", default="naver_상품등록_템플릿.xlsx")
    parser.add_argument("--dry-run", action="store_true", help="실제 API 호출 없이 생성될 JSON만 출력")
    parser.add_argument("--limit", type=int, default=None, help="처음 N개 상품만 처리 (실등록 테스트용)")
    parser.add_argument("--force", action="store_true", help="이미 등록 로그에 있는 상품도 다시 등록 (중복 생성 주의)")
    args = parser.parse_args()

    rows = read_rows(args.file)
    if not rows:
        print("등록할 상품이 없습니다. 엑셀 파일의 필수 항목을 확인하세요.")
        sys.exit(1)

    log = load_log()

    if not args.force:
        before = len(rows)
        rows = [r for r in rows if r["영문상품명"] not in log]
        skipped = before - len(rows)
        if skipped:
            print(f"이미 등록된 상품 {skipped}개는 건너뜁니다 (registered_log.json 참고, --force로 강제 재등록 가능).")

    if args.limit is not None:
        rows = rows[: args.limit]

    if not rows:
        print("처리할 상품이 없습니다 (모두 이미 등록됨).")
        sys.exit(0)

    print(f"{len(rows)}개 상품 처리 시작 (dry_run={args.dry_run})")

    token = None
    if not args.dry_run:
        token = get_bearer_token(config.CLIENT_ID, config.CLIENT_SECRET)

    for row in rows:
        name = row["영문상품명"]
        try:
            result = register_one(token, row, args.dry_run)
            if result and not args.dry_run:
                log[name] = result
                save_log(log)
        except Exception as e:  # noqa: BLE001
            print(f"  [오류] {name}: {e}")


if __name__ == "__main__":
    main()
